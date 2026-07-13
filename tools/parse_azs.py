#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сбор базы АЗС/АГЗС по региону для холодного обзвона.

Источники (все бесплатные, без платных подписок):
  - OpenStreetMap Overpass API  — точки заправок: название, бренд/оператор,
                                  телефон, адрес, сайт (без API-ключа).
  - Nominatim (OSM)             — геокодирование названия региона в границу.
  - ГИР БО (bo.nalog.ru)        — оборот (выручка) по ИНН из открытой
                                  бухотчётности ФНС (best-effort).
  - DaData (опционально)        — если задан DADATA_TOKEN, ищет ИНН по
                                  названию+городу, чтобы подтянуть оборот.
                                  Бесплатный тариф DaData: до 10 000 запросов/сут.

Использование:
    python3 tools/parse_azs.py "Краснодарский край"
    python3 tools/parse_azs.py                # спросит регион интерактивно

Результат: xlsx рядом со скриптом, напр. azs_krasnodarskij-kraj.xlsx
Колонки: Название | Тип | Бренд/Оператор | Телефон | Адрес | Сайт | ИНН |
         Оборот за год, ₽ | Год отчёта | Координаты | Источник
"""

import os
import re
import sys
import time
import json
import unicodedata

import requests

USER_AGENT = "syndicat-azs-leadgen/1.0 (sd-kt.ru; contact: nikoai2911@gmail.com)"
NOMINATIM = "https://nominatim.openstreetmap.org/search"
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
]
DADATA_TOKEN = os.environ.get("DADATA_TOKEN", "").strip()

SESSION = requests.Session()
SESSION.headers.update({"User-Agent": USER_AGENT})


# ---------------------------------------------------------------------------
# 1. Регион -> граница OSM
# ---------------------------------------------------------------------------
def find_region_area(region: str):
    """Возвращает (area_id, отображаемое_имя) для Overpass или None."""
    params = {
        "q": region,
        "format": "json",
        "limit": 5,
        "addressdetails": 1,
        "accept-language": "ru",
    }
    r = SESSION.get(NOMINATIM, params=params, timeout=60)
    r.raise_for_status()
    results = r.json()
    # Нужен полигон (relation) — область/край/город с границей.
    for item in results:
        if item.get("osm_type") == "relation":
            area_id = 3600000000 + int(item["osm_id"])
            return area_id, item.get("display_name", region)
    # Фолбэк: первый результат, если это way с границей
    for item in results:
        if item.get("osm_type") == "way":
            area_id = 2400000000 + int(item["osm_id"])
            return area_id, item.get("display_name", region)
    return None


# ---------------------------------------------------------------------------
# 2. Overpass -> все заправки в границе
# ---------------------------------------------------------------------------
def fetch_fuel_stations(area_id: int):
    query = f"""
    [out:json][timeout:240];
    area({area_id})->.a;
    (
      node["amenity"="fuel"](area.a);
      way["amenity"="fuel"](area.a);
      node["shop"="gas"](area.a);
      way["shop"="gas"](area.a);
    );
    out center tags;
    """
    last_err = None
    for url in OVERPASS_ENDPOINTS:
        try:
            r = SESSION.post(url, data={"data": query}, timeout=300)
            r.raise_for_status()
            return r.json().get("elements", [])
        except Exception as e:  # noqa: BLE001
            last_err = e
            print(f"  ! Overpass {url} не ответил ({e}), пробую зеркало...")
            time.sleep(3)
    raise RuntimeError(f"Все зеркала Overpass недоступны: {last_err}")


# ---------------------------------------------------------------------------
# 3. Разбор тегов одной точки
# ---------------------------------------------------------------------------
def pick_phone(tags: dict) -> str:
    for key in ("contact:phone", "phone", "contact:mobile", "contact:phone:mobile"):
        if tags.get(key):
            return tags[key].split(";")[0].strip()
    return ""


def pick_address(tags: dict, region_name: str, lat, lon) -> str:
    if tags.get("addr:full"):
        return tags["addr:full"]
    parts = []
    for key in ("addr:region", "addr:city", "addr:town", "addr:village",
                "addr:street", "addr:housenumber"):
        if tags.get(key):
            parts.append(tags[key])
    if parts:
        return ", ".join(parts)
    # Ничего нет — хотя бы регион + координаты для навигатора
    if lat and lon:
        return f"{region_name.split(',')[0]} (по координатам {lat:.5f}, {lon:.5f})"
    return ""


def classify(tags: dict) -> str:
    """АГЗС если есть газ (СУГ/КПГ), иначе АЗС; смешанные помечаем."""
    lpg = tags.get("fuel:lpg") == "yes" or tags.get("shop") == "gas"
    cng = tags.get("fuel:cng") == "yes" or tags.get("fuel:lng") == "yes"
    liquid = any(
        tags.get(k) == "yes"
        for k in ("fuel:diesel", "fuel:octane_92", "fuel:octane_95",
                  "fuel:octane_98", "fuel:octane_100", "fuel:petrol")
    )
    if (lpg or cng) and liquid:
        return "АЗС+АГЗС"
    if lpg or cng:
        return "АГЗС"
    return "АЗС"


def parse_element(el: dict, region_name: str) -> dict:
    tags = el.get("tags", {})
    lat = el.get("lat") or (el.get("center") or {}).get("lat")
    lon = el.get("lon") or (el.get("center") or {}).get("lon")
    name = (
        tags.get("name")
        or tags.get("brand")
        or tags.get("operator")
        or "Без названия"
    )
    inn = tags.get("inn") or tags.get("ref:INN") or ""
    return {
        "Название": name,
        "Тип": classify(tags),
        "Бренд/Оператор": tags.get("operator") or tags.get("brand") or "",
        "Телефон": pick_phone(tags),
        "Адрес": pick_address(tags, region_name, lat, lon),
        "Сайт": tags.get("website") or tags.get("contact:website") or "",
        "ИНН": inn,
        "Руководитель": "",
        "Оборот за год, ₽": "",
        "Год отчёта": "",
        "Координаты": f"{lat}, {lon}" if lat and lon else "",
        "Источник": "OpenStreetMap",
    }


# ---------------------------------------------------------------------------
# 4. Обогащение оборотом: ИНН -> ГИР БО
# ---------------------------------------------------------------------------
def dadata_find_party(name: str, city_hint: str = "") -> dict:
    """Ищет компанию по названию через DaData. Возвращает dict с полями
    inn / address / manager / phone (что доступно на бесплатном тарифе)."""
    if not DADATA_TOKEN:
        return {}
    try:
        url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/party"
        payload = {"query": f"{name} {city_hint}".strip(), "count": 1,
                   "status": ["ACTIVE"]}
        r = SESSION.post(
            url,
            headers={
                "Authorization": f"Token {DADATA_TOKEN}",
                "Content-Type": "application/json",
            },
            data=json.dumps(payload),
            timeout=30,
        )
        r.raise_for_status()
        sug = r.json().get("suggestions", [])
        if not sug:
            return {}
        d = sug[0]["data"]
        phones = d.get("phones") or []
        return {
            "inn": d.get("inn", ""),
            "address": (d.get("address") or {}).get("value", ""),
            "manager": (d.get("management") or {}).get("name", ""),
            "phone": (phones[0].get("value") if phones else ""),
        }
    except Exception:  # noqa: BLE001
        return {}


_GIRBO_READY = False


def _girbo_init():
    """Заводит сессию на ГИР БО (нужны куки из /nbo/context)."""
    global _GIRBO_READY
    if _GIRBO_READY:
        return
    try:
        SESSION.get("https://bo.nalog.gov.ru/nbo/context", timeout=30)
    except Exception:  # noqa: BLE001
        pass
    _GIRBO_READY = True


def girbo_revenue(inn: str):
    """Оборот (выручка) из ГИР БО по ИНН. Возвращает (сумма_в_рублях, год).

    Поиск на bo.nalog.gov.ru сразу отдаёт bfo.gainSum (выручка последнего
    сданного периода) в тыс. руб — переводим в рубли.
    """
    if not inn or not re.fullmatch(r"\d{10,12}", inn):
        return None, None
    _girbo_init()
    try:
        r = SESSION.get(
            "https://bo.nalog.gov.ru/advanced-search/organizations/search",
            params={"query": inn, "page": 0, "size": 1},
            headers={"Accept": "application/json"},
            timeout=30,
        )
        r.raise_for_status()
        content = r.json().get("content", [])
        if not content:
            return None, None
        bfo = content[0].get("bfo") or {}
        gain = bfo.get("gainSum")
        period = bfo.get("period")
        if gain is not None:
            return int(gain) * 1000, (int(period) if period else None)
    except Exception:  # noqa: BLE001
        return None, None
    return None, None


# ---------------------------------------------------------------------------
# 5. Выгрузка в Excel
# ---------------------------------------------------------------------------
def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    translit = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
        "ж": "zh", "з": "z", "и": "i", "й": "j", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya", " ": "-",
    }
    out = "".join(translit.get(ch.lower(), ch.lower()) for ch in text)
    out = re.sub(r"[^a-z0-9-]+", "", out)
    return re.sub(r"-+", "-", out).strip("-") or "region"


def write_xlsx(rows, region_slug):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cols = ["Название", "Тип", "Бренд/Оператор", "Телефон", "Адрес", "Сайт",
            "ИНН", "Руководитель", "Оборот за год, ₽", "Год отчёта",
            "Координаты", "Источник"]

    wb = Workbook()
    ws = wb.active
    ws.title = "АЗС и АГЗС"

    head_fill = PatternFill("solid", fgColor="1F3A5F")
    head_font = Font(bold=True, color="FFFFFF")
    for c, name in enumerate(cols, 1):
        cell = ws.cell(1, c, name)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(c, "") for c in cols])

    widths = [34, 10, 24, 20, 46, 30, 14, 28, 18, 11, 24, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       f"azs_{region_slug}.xlsx")
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    region = " ".join(a for a in sys.argv[1:] if not a.startswith("--")).strip()
    if not region:
        region = input("Введите регион (напр. «Краснодарский край» или город): ").strip()
    if not region:
        print("Регион не задан. Выход.")
        return

    print(f"\n▸ Ищу границы: {region}")
    area = find_region_area(region)
    if not area:
        print("  ✗ Не нашёл регион в OSM. Уточни название (как на карте).")
        return
    area_id, display = area
    print(f"  ✓ {display}")

    print("▸ Тяну заправки из OpenStreetMap...")
    elements = fetch_fuel_stations(area_id)
    print(f"  ✓ Найдено точек: {len(elements)}")

    rows = [parse_element(el, display) for el in elements]

    # Дедуп по (название + координаты)
    seen, unique = set(), []
    for r in rows:
        key = (r["Название"].lower(), r["Координаты"])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    rows = unique
    print(f"  ✓ После дедупа: {len(rows)}")

    with_phone = sum(1 for r in rows if r["Телефон"])
    print(f"  · С телефоном: {with_phone} / {len(rows)}")

    # Обогащение оборотом (best-effort, может занять время)
    enrich = "--no-revenue" not in sys.argv
    if enrich:
        print("▸ Подтягиваю оборот (ГИР БО)"
              + (" + ИНН через DaData" if DADATA_TOKEN else "") + "...")
        done = 0
        party_cache = {}   # ключ запроса -> dict DaData
        rev_cache = {}     # ИНН -> (оборот, год)
        total = len(rows)
        for i, r in enumerate(rows, 1):
            inn = r["ИНН"]
            if not inn and DADATA_TOKEN:
                city = ""
                m = re.search(r"(?:г\.?\s*)?([А-ЯЁ][а-яё-]+)", r["Адрес"])
                if m:
                    city = m.group(1)
                q = (r["Бренд/Оператор"] or r["Название"], city)
                if q in party_cache:
                    party = party_cache[q]
                else:
                    party = dadata_find_party(q[0], q[1])
                    party_cache[q] = party
                    time.sleep(0.15)
                inn = party.get("inn", "")
                r["ИНН"] = inn
                r["Руководитель"] = party.get("manager", "")
                if not r["Телефон"] and party.get("phone"):
                    r["Телефон"] = party["phone"]
                if party.get("address") and "координатам" in r["Адрес"]:
                    r["Адрес"] = party["address"]
            if inn:
                if inn in rev_cache:
                    rev, year = rev_cache[inn]
                else:
                    rev, year = girbo_revenue(inn)
                    rev_cache[inn] = (rev, year)
                    time.sleep(0.15)
                if rev is not None:
                    r["Оборот за год, ₽"] = rev
                    r["Год отчёта"] = year
                    done += 1
            if i % 100 == 0:
                print(f"    · обработано {i}/{total}, оборот у {done}")
        print(f"  ✓ Оборот найден у: {done}")
    else:
        print("▸ Оборот пропущен (--no-revenue)")

    out = write_xlsx(rows, slugify(region))
    print(f"\n✅ Готово: {out}")
    print(f"   Всего компаний: {len(rows)} | с телефоном: {with_phone}")
    if not DADATA_TOKEN:
        print("   💡 Хочешь больше оборотов/ИНН — заведи бесплатный токен DaData")
        print("      и запусти:  DADATA_TOKEN=твой_токен python3 tools/parse_azs.py \"…\"")


if __name__ == "__main__":
    main()
