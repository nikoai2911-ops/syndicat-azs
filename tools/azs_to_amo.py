#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Конвертер выгрузки парсера АЗС (tools/parse_azs*.py) в файл импорта amoCRM.

Вход:  xlsx парсера (Название, Тип, Бренд/Оператор, Телефон, Адрес, Сайт, ИНН,
       Руководитель, Оборот за год, Год отчёта, Координаты, Источник)
Выход: xlsx для импорта в amoCRM (Компании → ... → Импорт):
       одна строка = компания + контактный телефон + сделка в воронке
       «Холодный обзвон», этап «Взят в работу», ответственный Сазонов Никита.

Дедупликация по нормализованному телефону: у сетей один 8-800 на много точек —
адреса объединяются в одну компанию (в примечание).

Запуск:
    python3 tools/azs_to_amo.py tools/azs2gis_krasnoyarsk.xlsx "Красноярский край"
"""
import re
import sys

import openpyxl

RESPONSIBLE = "Сазонов Никита"
STAGE = "Взят в работу"          # этап воронки «Холодный обзвон»


def norm_phone(raw):
    """+7 391 290-60-85 → +73912906085; мусор → None."""
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 11 and digits[0] in "78":
        return "+7" + digits[1:]
    if len(digits) == 10:
        return "+7" + digits
    return None


def convert(src_path, region):
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    companies = {}  # phone -> dict
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        phone = norm_phone(row[idx["Телефон"]])
        if not phone:
            skipped += 1
            continue
        name = (row[idx["Название"]] or "").strip()
        brand = (row[idx.get("Бренд/Оператор", idx.get("Бренд"))] or "") if "Бренд/Оператор" in idx or "Бренд" in idx else ""
        brand = (brand or "").strip()
        typ = (row[idx["Тип"]] or "").strip()
        addr = (row[idx["Адрес"]] or "").strip()
        site = (row[idx["Сайт"]] or "").strip() if "Сайт" in idx else ""
        inn = str(row[idx["ИНН"]] or "").strip() if "ИНН" in idx else ""
        head = (row[idx["Руководитель"]] or "").strip() if "Руководитель" in idx else ""

        if not name or name.lower() in ("без названия", "агзс", "азс"):
            name = brand or (typ + " " + addr) or phone

        c = companies.setdefault(phone, {
            "name": name, "type": typ, "brand": brand, "addrs": [],
            "site": site, "inn": inn, "head": head,
        })
        if addr and addr not in c["addrs"]:
            c["addrs"].append(addr)
        # более информативное название побеждает
        if c["name"].startswith(("АГЗС", "АЗС")) and name and not name.startswith(("АГЗС", "АЗС")):
            c["name"] = name

    out = openpyxl.Workbook()
    o = out.active
    o.title = "import"
    o.append([
        "Название сделки", "Этап сделки", "Ответственный",
        "Название компании", "Рабочий телефон", "Адрес", "Web", "ИНН",
        "Примечание", "Теги",
    ])
    tag_region = region.lower()
    for phone, c in companies.items():
        note_parts = []
        if c["type"]:
            note_parts.append("Тип: " + c["type"])
        if c["brand"] and c["brand"] != c["name"]:
            note_parts.append("Бренд: " + c["brand"])
        if c["head"]:
            note_parts.append("Руководитель: " + c["head"])
        if len(c["addrs"]) > 1:
            note_parts.append("Точки (%d): %s" % (len(c["addrs"]), "; ".join(c["addrs"])))
        company_name = "%s (%s)" % (c["name"], region)
        o.append([
            "Обзвон: " + c["name"],
            STAGE,
            RESPONSIBLE,
            company_name,
            phone,
            c["addrs"][0] if c["addrs"] else "",
            c["site"],
            c["inn"],
            " | ".join(note_parts),
            "холодная база, %s, 2гис" % tag_region,
        ])

    dst = re.sub(r"\.xlsx$", "", src_path) + "_amo.xlsx"
    out.save(dst)
    print("компаний после дедупа: %d (строк пропущено без телефона: %d)" % (len(companies), skipped))
    print("файл:", dst)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("usage: azs_to_amo.py <файл_парсера.xlsx> <регион>")
    convert(sys.argv[1], sys.argv[2])
