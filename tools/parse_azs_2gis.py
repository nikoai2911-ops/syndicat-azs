#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сбор базы АЗС/АГЗС по ГОРОДУ из 2ГИС — с телефонами — в Excel.

Зачем: у 2ГИС телефоны есть почти у каждой заправки (в отличие от OSM, где
их 14–59%). Скрипт делает то же, что вы вручную в 2ГИС, только автоматически:
ищет АЗС и АГЗС по городу, разворачивает сети по филиалам и с карточки каждой
точки забирает название, телефон, адрес и координаты — сразу в таблицу.

Как это обходит бот-защиту 2ГИС:
    Используется НАСТОЯЩИЙ Chrome (channel="chrome"), а не встроенный в
    Playwright Chromium — 2ГИС пускает реальный браузер. Поэтому Chrome
    должен быть установлен, и окно будет видно во время работы (headful).

Установка (один раз):
    python3 -m pip install --user playwright openpyxl
    python3 -m playwright install chromium      # запасной движок

Использование:
    python3 tools/parse_azs_2gis.py krasnoyarsk
    python3 tools/parse_azs_2gis.py "Новосибирск"
    python3 tools/parse_azs_2gis.py moscow --workers 4
    python3 tools/parse_azs_2gis.py krasnoyarsk --no-branches   # быстро, без сетей
    python3 tools/parse_azs_2gis.py krasnoyarsk --limit 30      # проба на 30 точках

Город можно задавать 2ГИС-«слагом» из адресной строки (krasnoyarsk, moscow,
novosibirsk…) или по-русски — частые города распознаются, для остального
скрипт транслитерует и пробует; если не открылось — подставьте слаг из URL
2gis.ru/<слаг>/.

Результат: tools/azs2gis_<город>.xlsx
Колонки совпадают с parse_azs.py, поэтому файлы легко объединять.

После сохранения Excel база АВТОМАТИЧЕСКИ заливается в amoCRM (воронка
«Холодный обзвон → Взят в работу», ответственный Никита, задачи на завтра;
дубли по телефону пропускаются). Отключить: --no-amo. Метка региона для
тегов берётся из названия города, переопределить: --region "Красноярский край".
"""

import argparse
import asyncio
import os
import re
import sys

# ---------------------------------------------------------------------------
# Город -> 2ГИС-слаг
# ---------------------------------------------------------------------------
CITY_SLUGS = {
    "москва": "moscow", "санкт-петербург": "spb", "петербург": "spb",
    "спб": "spb", "новосибирск": "novosibirsk", "екатеринбург": "ekaterinburg",
    "казань": "kazan", "нижний новгород": "n_novgorod", "челябинск": "chelyabinsk",
    "самара": "samara", "омск": "omsk", "ростов-на-дону": "rostov",
    "ростов": "rostov", "уфа": "ufa", "красноярск": "krasnoyarsk",
    "воронеж": "voronezh", "пермь": "perm", "волгоград": "volgograd",
    "краснодар": "krasnodar", "саратов": "saratov", "тюмень": "tyumen",
    "тольятти": "tolyatti", "ижевск": "izhevsk", "барнаул": "barnaul",
    "иркутск": "irkutsk", "хабаровск": "khabarovsk", "владивосток": "vladivostok",
    "ярославль": "yaroslavl", "махачкала": "mahachkala", "томск": "tomsk",
    "оренбург": "orenburg", "кемерово": "kemerovo", "новокузнецк": "novokuznetsk",
    "рязань": "ryazan", "астрахань": "astrahan", "пенза": "penza",
    "липецк": "lipetsk", "тула": "tula", "киров": "kirov", "чебоксары": "cheboksary",
    "калининград": "kaliningrad", "брянск": "bryansk", "курск": "kursk",
    "сочи": "sochi", "ставрополь": "stavropol", "улан-удэ": "ulan-ude",
    "тверь": "tver", "магнитогорск": "magnitogorsk", "иваново": "ivanovo",
    "сургут": "surgut", "белгород": "belgorod", "владимир": "vladimir",
    "нижний тагил": "nizhny_tagil", "чита": "chita", "архангельск": "arhangelsk",
    "симферополь": "simferopol", "смоленск": "smolensk", "курган": "kurgan",
    "орёл": "oryol", "орел": "oryol", "вологда": "vologda", "якутск": "yakutsk",
    "владикавказ": "vladikavkaz", "мурманск": "murmansk", "тамбов": "tambov",
    "грозный": "groznyy", "стерлитамак": "sterlitamak", "кострома": "kostroma",
    "петрозаводск": "petrozavodsk", "нижневартовск": "nizhnevartovsk",
    "новороссийск": "novorossiysk", "йошкар-ола": "yoshkar-ola",
    "абакан": "abakan", "сыктывкар": "syktyvkar", "норильск": "norilsk",
}

TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
    "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    " ": "-", "-": "-",
}


def slugify_city(city: str) -> str:
    key = city.strip().lower()
    if key in CITY_SLUGS:
        return CITY_SLUGS[key]
    if re.fullmatch(r"[a-z0-9_\-]+", key):
        return key  # уже слаг
    out = "".join(TRANSLIT.get(ch, ch) for ch in key)
    out = re.sub(r"[^a-z0-9_\-]+", "", out)
    return out.strip("-") or "moscow"


def region_label(city_input: str, slug: str) -> str:
    """Человекочитаемая метка региона для тегов amoCRM."""
    key = (city_input or "").strip()
    if key and not re.fullmatch(r"[a-z0-9_\-]+", key.lower()):
        return key[0].upper() + key[1:]          # ввели по-русски
    for rus, s in CITY_SLUGS.items():            # ввели слаг — ищем русское имя
        if s == slug:
            return rus[0].upper() + rus[1:]
    return slug


# ---------------------------------------------------------------------------
# Разбор одной карточки фирмы
# ---------------------------------------------------------------------------
def normalize_phone(raw: str) -> str:
    d = re.sub(r"\D", "", raw)
    if len(d) == 11 and d[0] == "8":
        d = "7" + d[1:]
    if len(d) == 11 and d[0] == "7":
        return f"+7 {d[1:4]} {d[4:7]}-{d[7:9]}-{d[9:11]}"
    if len(d) == 10:  # без кода страны
        return f"+7 {d[0:3]} {d[3:6]}-{d[6:8]}-{d[8:10]}"
    return raw.strip()


def classify(queries: set, name: str, page_text: str) -> str:
    t = (name + " " + page_text[:600]).lower()
    gas = ("агзс" in {q.lower() for q in queries}
           or "газ" in t or "суг" in t or "пропан" in t or "метан" in t)
    petrol = ("азс" in {q.lower() for q in queries} and "агзс" not in {q.lower() for q in queries})
    if gas and petrol:
        return "АЗС+АГЗС"
    if gas:
        return "АГЗС"
    return "АЗС"


async def parse_firm(page, slug: str, firm_id: str, queries: set) -> dict:
    url = f"https://2gis.ru/{slug}/firm/{firm_id}"
    await page.goto(url, wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(1200)
    html = await page.content()
    if "Forbidden" in html[:1500]:
        raise RuntimeError("Forbidden")

    # имя
    name = ""
    h1 = page.locator("h1").first
    try:
        if await h1.count():
            name = (await h1.inner_text()).replace("\xa0", " ").strip()
    except Exception:
        pass

    # телефоны — из href tel:
    tels = []
    for raw in re.findall(r'href="tel:([+0-9]+)"', html):
        ph = normalize_phone(raw)
        if ph not in tels:
            tels.append(ph)

    # координаты — из ссылки построения маршрута points/%7C<lon>%2C<lat>
    lat = lon = ""
    m = re.search(r"points/%7C([0-9.]+)%2C([0-9.]+)", html)
    if m:
        lon, lat = m.group(1), m.group(2)

    # адрес — ссылка на /geo/
    addr = ""
    try:
        geo = page.locator('a[href*="/geo/"]').first
        if await geo.count():
            addr = (await geo.inner_text()).replace("\xa0", " ").strip()
    except Exception:
        pass
    if not addr:
        # из <title>: «Имя, рубрика, АДРЕС, Город — 2ГИС»
        t = await page.title()
        parts = [p.strip() for p in t.split("—")[0].split(",")]
        if len(parts) >= 3:
            addr = ", ".join(parts[2:]).strip()

    # сайт — ссылка-редирект link.2gis.ru/...?<реальный_url> или прямой домен
    site = ""
    m2 = re.search(r'href="https?://link\.2gis\.ru/[^"]*\?(https?://[^"&]+)"', html)
    if m2:
        site = m2.group(1)

    page_text = ""
    try:
        page_text = await page.inner_text("body")
    except Exception:
        pass

    brand = name
    # бренд сети из заголовка (до запятой), напр. «Лукойл»
    return {
        "Название": name or "Без названия",
        "Тип": classify(queries, name, page_text),
        "Бренд/Оператор": brand,
        "Телефон": "; ".join(tels),
        "Адрес": addr,
        "Сайт": site,
        "ИНН": "",
        "Руководитель": "",
        "Оборот за год, ₽": "",
        "Год отчёта": "",
        "Координаты": f"{lat}, {lon}" if lat and lon else "",
        "Источник": "2ГИС",
    }


# ---------------------------------------------------------------------------
# Сбор ссылок с выдачи и страниц сетей
# ---------------------------------------------------------------------------
async def collect_links(page):
    firms = await page.eval_on_selector_all(
        'a[href*="/firm/"]', "els=>els.map(e=>e.getAttribute('href'))")
    brs = await page.eval_on_selector_all(
        'a[href*="/branches/"]', "els=>els.map(e=>e.getAttribute('href'))")
    fids = set(m.group(1) for h in firms if (m := re.search(r"/firm/(\d+)", h)))
    bids = set(m.group(1) for h in brs if (m := re.search(r"/branches/(\d+)", h)))
    return fids, bids


async def collect_search(page, slug: str, query: str, max_pages: int = 30):
    """Постранично собирает id карточек и id сетей по одному запросу."""
    firm_ids, chain_ids = set(), set()
    empty_streak = 0
    for pg in range(1, max_pages + 1):
        url = f"https://2gis.ru/{slug}/search/{query}/page/{pg}"
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
        except Exception:
            break
        await page.wait_for_timeout(1800)
        if "Forbidden" in (await page.content())[:1500]:
            print(f"    ! страница {pg}: бот-блокировка, пауза…")
            await page.wait_for_timeout(5000)
            continue
        f, b = await collect_links(page)
        new = len(f - firm_ids) + len(b - chain_ids)
        firm_ids |= f
        chain_ids |= b
        print(f"    поиск «{query}» стр.{pg}: карточек={len(f)} сетей={len(b)} новых={new}")
        if new == 0:
            empty_streak += 1
            if empty_streak >= 2:
                break
        else:
            empty_streak = 0
    return firm_ids, chain_ids


async def expand_chain(page, slug: str, chain_id: str, max_scroll: int = 40):
    """Разворачивает сеть: собирает id всех филиалов со страницы /branches/."""
    url = f"https://2gis.ru/{slug}/branches/{chain_id}"
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)
    except Exception:
        return set()
    await page.wait_for_timeout(2000)
    if "Forbidden" in (await page.content())[:1500]:
        return set()
    prev = -1
    for _ in range(max_scroll):
        f, _b = await collect_links(page)
        if len(f) == prev:
            break
        prev = len(f)
        await page.mouse.move(350, 500)
        await page.mouse.wheel(0, 4000)
        await page.wait_for_timeout(800)
    f, _b = await collect_links(page)
    return f


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
COLS = ["Название", "Тип", "Бренд/Оператор", "Телефон", "Адрес", "Сайт",
        "ИНН", "Руководитель", "Оборот за год, ₽", "Год отчёта",
        "Координаты", "Источник"]


def write_xlsx(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "АЗС и АГЗС (2ГИС)"
    head_fill = PatternFill("solid", fgColor="1F3A5F")
    head_font = Font(bold=True, color="FFFFFF")
    for c, name in enumerate(COLS, 1):
        cell = ws.cell(1, c, name)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append([row.get(c, "") for c in COLS])
    widths = [34, 10, 24, 22, 46, 30, 14, 28, 18, 11, 24, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# main (async)
# ---------------------------------------------------------------------------
async def run(args):
    from playwright.async_api import async_playwright

    slug = slugify_city(args.city)
    queries = [q.strip() for q in args.queries.split(",") if q.strip()]
    profile_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), ".2gis_profile")

    print(f"▸ Город: {args.city}  ->  2gis.ru/{slug}")
    print(f"▸ Запросы: {', '.join(queries)}")

    async with async_playwright() as p:
        launch_kwargs = dict(
            user_data_dir=profile_dir, channel="chrome",
            headless=args.headless, locale="ru-RU",
            viewport={"width": 1400, "height": 950},
            args=["--disable-blink-features=AutomationControlled"],
        )
        try:
            ctx = await p.chromium.launch_persistent_context(**launch_kwargs)
        except Exception as e:
            print(f"  ! Не удалось запустить настоящий Chrome ({e}).")
            print("    Убедитесь, что Google Chrome установлен.")
            return
        page = ctx.pages[0] if ctx.pages else await ctx.new_page()

        # прогрев: заходим на главную города, чтобы получить куки 2ГИС
        try:
            await page.goto(f"https://2gis.ru/{slug}", wait_until="domcontentloaded",
                            timeout=60000)
            await page.wait_for_timeout(2500)
        except Exception:
            pass
        if "Forbidden" in (await page.content())[:1500]:
            print("  ! 2ГИС отдаёт «Forbidden» даже настоящему Chrome.")
            print("    Обычно это из-за IP (VPN/зарубежный адрес). Отключите VPN")
            print("    и попробуйте снова — с российского IP 2ГИС пускает.")
            await ctx.close()
            return

        # 1) собрать id карточек и сетей по каждому запросу
        firm_queries = {}      # firm_id -> set(query)
        chain_queries = {}     # chain_id -> set(query)
        for q in queries:
            fids, cids = await collect_search(page, slug, q)
            for fid in fids:
                firm_queries.setdefault(fid, set()).add(q)
            for cid in cids:
                chain_queries.setdefault(cid, set()).add(q)
        print(f"  ✓ карточек: {len(firm_queries)} | сетей: {len(chain_queries)}")

        # 2) развернуть сети в филиалы (если не отключено)
        if not args.no_branches and chain_queries:
            print("▸ Разворачиваю сети по филиалам…")
            for i, (cid, qs) in enumerate(list(chain_queries.items()), 1):
                fids = await expand_chain(page, slug, cid)
                for fid in fids:
                    firm_queries.setdefault(fid, set()).update(qs)
                print(f"    сеть {i}/{len(chain_queries)} (id {cid}): +{len(fids)} филиалов")
        print(f"  ✓ всего уникальных точек к обходу: {len(firm_queries)}")

        firm_ids = list(firm_queries)
        if args.limit:
            firm_ids = firm_ids[: args.limit]
            print(f"  · ограничение --limit: {len(firm_ids)}")

        # 3) обойти карточки (с небольшой параллельностью)
        print(f"▸ Собираю карточки (потоков: {args.workers})…")
        rows = []
        sem = asyncio.Semaphore(args.workers)
        pages_pool = [page] + [await ctx.new_page() for _ in range(args.workers - 1)]
        lock = asyncio.Lock()
        done = 0

        async def worker(idx, fid, wpage):
            nonlocal done
            async with sem:
                try:
                    row = await parse_firm(wpage, slug, fid, firm_queries[fid])
                except Exception as e:
                    row = None
                    if "Forbidden" in str(e):
                        await wpage.wait_for_timeout(4000)
                async with lock:
                    done += 1
                    if row:
                        rows.append(row)
                    if done % 20 == 0:
                        print(f"    · {done}/{len(firm_ids)}  (с телефоном "
                              f"{sum(1 for r in rows if r['Телефон'])})")

        tasks = []
        for i, fid in enumerate(firm_ids):
            wpage = pages_pool[i % len(pages_pool)]
            tasks.append(worker(i, fid, wpage))
        await asyncio.gather(*tasks)

        await ctx.close()

    # дедуп по (название + координаты) и по телефону-адресу
    seen, unique = set(), []
    for r in rows:
        key = (r["Название"].lower(), r["Координаты"] or r["Адрес"].lower())
        if key in seen:
            continue
        seen.add(key)
        unique.append(r)
    unique.sort(key=lambda r: (r["Тип"], r["Название"].lower()))

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       f"azs2gis_{slug}.xlsx")
    write_xlsx(unique, out)
    with_phone = sum(1 for r in unique if r["Телефон"])
    print(f"\n✅ Готово: {out}")
    print(f"   Точек: {len(unique)} | с телефоном: {with_phone} "
          f"({100*with_phone//max(len(unique),1)}%)")

    if not args.no_amo:
        region = args.region or region_label(args.city, slug)
        print(f"\n→ Заливаю в amoCRM (воронка «Холодный обзвон», регион «{region}»)…")
        try:
            from azs_push_amo import push
            push(out, region)
        except Exception as e:
            print("⚠ Не залилось в amoCRM:", e)
            print(f"  Файл сохранён. Повторить вручную: "
                  f"python3 tools/azs_push_amo.py {out} \"{region}\"")


def main():
    ap = argparse.ArgumentParser(
        description="База АЗС/АГЗС из 2ГИС с телефонами -> Excel")
    ap.add_argument("city", nargs="?", help="Город (слаг 2ГИС или по-русски)")
    ap.add_argument("--queries", default="АЗС,АГЗС",
                    help="Поисковые запросы через запятую (по умолч. АЗС,АГЗС)")
    ap.add_argument("--workers", type=int, default=3,
                    help="Сколько карточек грузить параллельно (по умолч. 3)")
    ap.add_argument("--no-branches", action="store_true",
                    help="Не разворачивать сети по филиалам (быстро, но меньше точек)")
    ap.add_argument("--limit", type=int, default=0,
                    help="Обойти не более N точек (для пробы)")
    ap.add_argument("--headless", action="store_true",
                    help="Без окна (2ГИС может заблокировать — не рекомендуется)")
    ap.add_argument("--no-amo", action="store_true",
                    help="Только Excel, не заливать в amoCRM")
    ap.add_argument("--region", default="",
                    help="Метка региона для тегов в amoCRM (по умолч. — из города)")
    args = ap.parse_args()

    if not args.city:
        args.city = input("Город (напр. krasnoyarsk или «Новосибирск»): ").strip()
    if not args.city:
        print("Город не задан. Выход.")
        return
    try:
        asyncio.run(run(args))
    except KeyboardInterrupt:
        print("\nПрервано пользователем.")


if __name__ == "__main__":
    main()
