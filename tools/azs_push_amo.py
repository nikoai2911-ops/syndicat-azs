#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Автозагрузка холодной базы АЗС в amoCRM (без браузера и импортов).

Берёт выгрузку парсера (tools/parse_azs*.py), нормализует и дедуплицирует
телефоны, пропускает номера, уже существующие в amo, и для каждого нового:
  компания (телефон, адрес, web) + примечание с деталями
  + сделка в «Холодный обзвон → Взят в работу» на Сазонова Никиту
  + задача «Связаться» на завтра 18:00
  + теги: холодная база, <регион>

Запуск:
    python3 tools/azs_push_amo.py tools/azs2gis_krasnoyarsk.xlsx "Красноярский край"
"""
import datetime
import re
import sys
import time

from amo_api import (PIPELINE_COLD, STATUS_COLD_NEW, USER_NIKITA, get, post)


def norm_phone(raw):
    if not raw:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 11 and digits[0] in "78":
        return "+7" + digits[1:]
    if len(digits) == 10:
        return "+7" + digits
    return None


def load_companies(src_path, region):
    import openpyxl
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    brand_col = "Бренд/Оператор" if "Бренд/Оператор" in idx else ("Бренд" if "Бренд" in idx else None)

    companies = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        phone = norm_phone(row[idx["Телефон"]])
        if not phone:
            continue
        name = (row[idx["Название"]] or "").strip()
        brand = (row[idx[brand_col]] or "").strip() if brand_col else ""
        typ = (row[idx["Тип"]] or "").strip()
        addr = (row[idx["Адрес"]] or "").strip()
        site = (str(row[idx["Сайт"]] or "")).strip() if "Сайт" in idx else ""
        head = (str(row[idx["Руководитель"]] or "")).strip() if "Руководитель" in idx else ""
        if not name or name.lower() in ("без названия", "агзс", "азс"):
            name = brand or (typ + " " + addr).strip() or phone
        c = companies.setdefault(phone, {
            "name": name, "type": typ, "brand": brand, "addrs": [],
            "site": site, "head": head,
        })
        if addr and addr not in c["addrs"]:
            c["addrs"].append(addr)
        if c["name"].startswith(("АГЗС", "АЗС")) and name and not name.startswith(("АГЗС", "АЗС")):
            c["name"] = name
    return companies


def phone_exists(phone):
    """Ищем номер среди контактов и компаний (по последним 10 цифрам)."""
    q = phone[-10:]
    for entity in ("contacts", "companies"):
        try:
            d = get("/%s?query=%s&limit=1" % (entity, q))
        except RuntimeError as e:
            if "-> 204" in str(e):
                continue
            raise
        if (d.get("_embedded") or {}).get(entity):
            return True
        time.sleep(0.15)
    return False


def tomorrow_ts():
    t = datetime.datetime.now().replace(hour=18, minute=0, second=0) + datetime.timedelta(days=1)
    return int(t.timestamp())


def create_entry(phone, c, region, due_ts):
    """Компания + примечание + сделка в холодной воронке + задача. Возвращает id сделки."""
    company_payload = [{
        "name": "%s (%s)" % (c["name"], region),
        "responsible_user_id": USER_NIKITA,
        "custom_fields_values": [
            {"field_code": "PHONE", "values": [{"value": phone, "enum_code": "WORK"}]},
        ] + ([{"field_code": "WEB", "values": [{"value": c["site"]}]}] if c.get("site") else [])
          + ([{"field_code": "ADDRESS", "values": [{"value": c["addrs"][0]}]}] if c.get("addrs") else []),
    }]
    comp = post("/companies", company_payload)
    comp_id = comp["_embedded"]["companies"][0]["id"]

    note_parts = []
    if c.get("type"):
        note_parts.append("Тип: " + c["type"])
    if c.get("brand") and c["brand"] != c["name"]:
        note_parts.append("Бренд: " + c["brand"])
    if c.get("head"):
        note_parts.append("Руководитель: " + c["head"])
    if len(c.get("addrs") or []) > 1:
        note_parts.append("Точки (%d): %s" % (len(c["addrs"]), "; ".join(c["addrs"])))
    if c.get("comment"):
        note_parts.append(c["comment"])
    if note_parts:
        post("/companies/notes", [{"entity_id": comp_id, "note_type": "common",
                                   "params": {"text": " | ".join(note_parts)}}])

    lead = post("/leads", [{
        "name": "Обзвон: " + c["name"],
        "pipeline_id": PIPELINE_COLD,
        "status_id": STATUS_COLD_NEW,
        "responsible_user_id": USER_NIKITA,
        "_embedded": {
            "companies": [{"id": comp_id}],
            "tags": [{"name": "холодная база"}, {"name": region.lower()}],
        },
    }])
    lead_id = lead["_embedded"]["leads"][0]["id"]

    post("/tasks", [{
        "task_type_id": 1,
        "text": "Первый звонок по базе: представиться, выявить потребность в запчастях АЗС/АГЗС",
        "complete_till": due_ts,
        "entity_id": lead_id,
        "entity_type": "leads",
        "responsible_user_id": USER_NIKITA,
    }])
    return lead_id


def push(src_path, region):
    companies = load_companies(src_path, region)
    print("в файле после дедупа: %d организаций" % len(companies))

    due_ts = tomorrow_ts()
    created = skipped = 0
    for phone, c in companies.items():
        if phone_exists(phone):
            skipped += 1
            print("  пропуск (уже в amo): %s %s" % (c["name"], phone))
            continue
        lead_id = create_entry(phone, c, region, due_ts)
        created += 1
        print("  создано: %s %s (сделка %d)" % (c["name"], phone, lead_id))
        time.sleep(0.25)

    print("итого: создано %d, пропущено как дубли %d" % (created, skipped))


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("usage: azs_push_amo.py <файл_парсера.xlsx> <регион>")
    push(sys.argv[1], sys.argv[2])
